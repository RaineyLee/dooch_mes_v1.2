import os
import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QBrush, QColor
from PyQt5 import uic, QtWidgets
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
# 차트 생성용
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

# 절대경로를 상대경로로 변경 하는 함수
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

#UI파일 연결
# main_window= uic.loadUiType(resource_path("/Users/black/projects/make_erp/main_window.ui"))[0] # Mac 사용시 ui 주소
main_window= uic.loadUiType(resource_path("./ui/overtime_chart_window.ui"))[0] # Window 사용시 ui 주소

#화면을 띄우는데 사용되는 Class 선언
class MainWindow(QWidget, main_window) :
    def __init__(self) :
        super().__init__()

        self.setupUi(self)
        self.setWindowTitle("잔업시간 조회/통계")

        self.canvas_bar = None
        self.canvas_pie = None

        self.monthly_dept_report()
        self.monthly_sum_report()

        self.layout_setting()
        self.slots()

    def slots(self):
        pass

    def layout_setting(self):        
     # 전체 레이아웃
        self.main_layout = QVBoxLayout()  # 세로로 정렬
        self.table_layout = QVBoxLayout()
        self.chart_layout = QHBoxLayout()
        self.table_layout.addWidget(self.tbl_dept_info)
        self.chart_layout.addWidget(self.canvas_bar)
        self.chart_layout.addWidget(self.canvas_pie)
        
        # 메인 레이아웃에 추가
        self.main_layout.addLayout(self.table_layout)
        self.main_layout.setAlignment(Qt.AlignTop)  # 위쪽 정렬
        self.main_layout.addLayout(self.chart_layout)

        self.setLayout(self.main_layout)

    def make_chart(self, column_name, result):
        plt.rc('font', family='Malgun Gothic')
        # Check and remove existing canvas if it exists
        if self.canvas_bar:
            self.chart_layout.removeWidget(self.canvas_bar)
            self.canvas_bar.deleteLater()
        
        year_month = column_name
        overtime = result
        
        # Create a new figure and canvas
        fig_bar = plt.Figure()
        self.canvas_bar = FigureCanvas(fig_bar) # pyqt5에서 matplotlib 사용하기 위해 FigureCanvas 객체 생성
        self.chart_layout.addWidget(self.canvas_bar)

        # Extracting data
        year_month = column_name[1:13]  # Assuming column_name includes a '날짜' column
        overtime = result[0][1:13]      # Assuming result is a list of lists with the overtime data

        # barchart Plotting the data
        self.ax_bar = fig_bar.add_subplot(111)
        self.bars = self.ax_bar.bar(year_month, overtime)
        self.ax_bar.set_title('월별 잔업시간')
        self.ax_bar.set_xlabel('월')
        self.ax_bar.set_ylabel('잔업시간')

        # Redraw the canvas
        self.canvas_bar.draw()

        self.canvas_bar.mpl_connect('button_press_event', self.on_click)

    def on_click(self, event):
        if event.inaxes == self.ax_bar and self.bars is not None:
            for bar in self.bars:
                if bar.contains(event)[0]:
                    # label = bar.get_x() + bar.get_width() / 2
                    # value = bar.get_height()
                    col = bar.get_x() + bar.get_width() / 2
                    month = int(col) + 1
                    
                    result = self.on_click_table_info(col)
                    label = result[0]
                    value = result[1]

                    self.show_pie_chart(label, value, month)
                    break
    
    def on_click_table_info(self, arg):
        row = self.tbl_dept_info.rowCount()
        col = int(arg) + 1

        list_value = [] 
        for i in range(row):
            value = self.tbl_dept_info.item(i,col)
            list_value.append(value.text())
        list_value = list(map(float, list_value))

        list_dept = [] 
        for i in range(row):
            dept = self.tbl_dept_info.item(i,0)
            list_dept.append(dept.text())

        return list_dept, list_value

    def show_pie_chart(self, label, value, col):
        plt.rc('font', family='Malgun Gothic')
        
        if self.canvas_pie:
            self.chart_layout.removeWidget(self.canvas_pie)
            self.canvas_pie.deleteLater()
        
        fig_pie = plt.Figure()
        self.canvas_pie = FigureCanvas(fig_pie)
        self.chart_layout.addWidget(self.canvas_pie)


         # 파이 차트 생성
        ax_pie = fig_pie.add_subplot(111)
        # values = [value, float(100) - value]  # value와 나머지 비율 계산 (예: 100에서 value를 뺀 값)
        # labels = [label, 'Others']     # 항목과 나머지 항목의 레이블 설정
        ax_pie.pie(value, labels=label, autopct='%1.1f%%')  # 파이 차트 그리기
        ax_pie.set_title(f'{col}월 부서별 잔업시간')  # 차트 제목 설정

        self.canvas_pie.draw()  # 캔버스 갱신
        
    def refresh_report(self):
        option = QtWidgets.QMessageBox.question(self, "QMessageBox", f"잔업 정보를 새로고침 하시겠습니까?", 
                                        QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.Yes)
            
        if option == QtWidgets.QMessageBox.Cancel:
            return
        elif option == QtWidgets.QMessageBox.No:
            return
        elif option == QtWidgets.QMessageBox.Yes: 
            self.monthly_dept_report()
            self.monthly_sum_report()

    def monthly_dept_report(self):
        self.tbl_dept_info.show()
        
        from db.db_select import Select
        select = Select()
        result, column_names = select.select_dept_monthly()

        self.make_dept_table(len(result), result, column_names)

    def monthly_sum_report(self):
        from db.db_select import Select
        select = Select()
        result, column_names = select.select_monthly_sum()

        self.make_chart(column_names, result)

    def make_dept_table(self, num, arr_1, column_names):   
        self.tbl_dept_info.setRowCount(0) # clear()는 행은 그대로 내용만 삭제, 행을 "0" 호출 한다.

        col = len(column_names)

        self.tbl_dept_info.setRowCount(num)
        self.tbl_dept_info.setColumnCount(col)
        self.tbl_dept_info.setHorizontalHeaderLabels(column_names)

        for i in range(num):
            for j in range(col): # 아니면 10개
                cell_value = arr_1[i][j]

                item = QTableWidgetItem(str(cell_value))
                item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)

                # 셀 값이 0인 경우 문자 색을 흰색으로 설정
                if cell_value == 0:
                    item.setForeground(QBrush(QColor(255, 255, 255)))

                self.tbl_dept_info.setItem(i, j, item)

        # 컨텐츠의 길이에 맞추어 컬럼의 길이를 자동으로 조절
        ################################################################
        table = self.tbl_dept_info
        header = table.horizontalHeader()

        # QSS 스타일 적용 (헤더 배경 색을 연한 회색으로 변경)
        table.setStyleSheet("""
            QHeaderView::section {
                background-color: lightgray;
                color: black;
                border: 1px solid #d6d6d6;
            }
        """)

        for i in range(col):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        ################################################################

        # 테이블의 길이에 맞추어 컬럼 길이를 균등하게 확장
        self.tbl_dept_info.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    
     # 테이블에 남겨진 정보를 엑셀로 변환
    def make_file(self):
        rows_dept_table = self.tbl_dept_info.rowCount()
        cols_dept_table = self.tbl_dept_info.columnCount()

        list_dept_1 = [] # 최종적으로 사용할 리스트는 for문 밖에 선언
        for i in range(rows_dept_table):
            list_dept_2 = [] # 2번째 for문 안쪽에서 사용할 리스트 선언
            for j in range(cols_dept_table): 
                data_dept = self.tbl_dept_info.item(i,j)
                list_dept_2.append(data_dept.text())
            list_dept_1.append(list_dept_2)

        num_dept = len(list_dept_1)

        self.make_excel(list_dept_1, num_dept)

    # 엑셀 파일을 만들고 넘겨진 배열 정보를 이용하여 sheet에 정보를 기입/저장 함.
    def make_excel(self, list_dept_1, num_dept):
        self.msg_box("자료저장", "부서 잔업정보가 생성 됩니다.")

        wb = openpyxl.Workbook()
        wb.create_sheet(index=0, title='부서잔업정보')

        dept_sheet = wb['부서잔업정보']

        column_count = self.tbl_dept_info.columnCount()
        dept_headers = []
        for col in range(column_count):
            header_item = self.tbl_dept_info.horizontalHeaderItem(col)
            if header_item:
                dept_headers.append(header_item.text())

        dept_sheet.append(dept_headers)

        for i in range(num_dept):
            for j in range(len(dept_headers)):
                dept_sheet.cell(row=i+2, column=j+1, value=list_dept_1[i][j])
        
        ## 각 칼럼에 대해서 모든 셀값의 문자열 개수에서 1.1만큼 곱한 것들 중 최대값을 계산한다.
        for column_cells in dept_sheet.columns:
            # length = max(len(str(cell.value))*1.1 for cell in column_cells)
            dept_sheet.column_dimensions[column_cells[0].column_letter].width = 20
            ## 셀 가운데 정렬
            for cell in dept_sheet[column_cells[0].column_letter]:
                cell.alignment = Alignment(horizontal='center')
        
        fname = self.file_save()

        try:
            if fname:
                self.save_excel(wb, fname)
        except Exception as e:
            self.msg_box("Error", str(e))

    # 파일 저장 대화상자(파일명 만들기)
    def file_save(self):
        now = datetime.now()
        arg_1 = now.strftime('%Y-%m-%d %H-%M-%S')
        adress = "./excel/download_" + arg_1 + ".xlsx"

        dialog = QFileDialog(self)
        qurl  = dialog.getSaveFileName(parent=self, caption='Save file', directory=adress)
        
        url = qurl[0]
        try:
            return url
        except Exception as e:
            QMessageBox.about(self, 'Warning', e)

    def save_excel(self, workbook, file_name):
        workbook.save(file_name)

    def select_all(self):
        import total_overtime as total_overtime_window

        self.total_window = total_overtime_window.MainWindow()
        self.total_window.show()

    def select_dept(self):
        import dept_overtime as select_dept_window

        self.dept_window = select_dept_window.DeptMainWindow()
        self.dept_window.show()
    
    def select_emp(self):
        import emp_overtime as select_emp_window

        self.emp_window = select_emp_window.MainWindow()
        self.emp_window.show() 

    # def select_month(self):
    #     import emp_overtime_month as select_emp_month_window

    #     self.emp_month_window = select_emp_month_window.MainWindow()
    #     self.emp_month_window.show()     
    
    def update_emp(self):
        import emp_overtime_update as update_emp_window

        self.emp_update_window = update_emp_window.MainWindow()
        self.emp_update_window.show() 

    def input_emp(self):
        import emp_overtime_input as input_emp_window

        self.emp_input_window = input_emp_window.MainWindow()
        self.emp_input_window.show() 

    def upload_overtime(self):
        import upload as upload_window

        self.upload_window = upload_window.MainWindow()
        self.upload_window.show()

    def emp_master(self):
        import emp_info as emp_info

        self.emp_master = emp_info.MainWindow()
        self.emp_master.show()

    def window_close(self):
        self.close()

    # def upload_location(self):        
    #     import upload_location as inv_loc

    #     self.location = inv_loc.WindowClass() #메인창에서 띄우려면 메인창을 뜻하는 self 추가
    #     self.location.show() #메인창에서 띄우려면 메인창을 뜻하는 self 추가

    # def upload_barcode(self):
    #     import upload_barcode as bar_loc

    #     self.barcode = bar_loc.WindowClass() #메인창에서 띄우려면 메인창을 뜻하는 self 추가
    #     self.barcode.show() #메인창에서 띄우려면 메인창을 뜻하는 self 추가

    # def upload_saleslist(self):
    #     import upload_saleslist as saleslist

    #     self.saleslist = saleslist.WindowClass() #메인창에서 띄우려면 메인창을 뜻하는 self 추가
    #     self.saleslist.show() #메인창에서 띄우려면 메인창을 뜻하는 self 추가

    # def item_location(self):
    #     import toexcel_location as item_loc

    #     self.item_loc = item_loc.WindowClass() #메인창에서 띄우려면 메인창을 뜻하는 self 추가
    #     self.item_loc.show() #메인창에서 띄우려면 메인창을 뜻하는 self 추가

    # def make_cjnumber(self):
    #     import CJ_number_v1_2 as cj_number

    #     self.cj_number = cj_number.WindowClass() #메인창에서 띄우려면 메인창을 뜻하는 self 추가
    #  self.cj_number.show() #메인창에서 띄우려면 메인창을 뜻하는 self 추가

    def msg_box(self, arg_1, arg_2):
        msg = QMessageBox()
        msg.setWindowTitle(arg_1)               # 제목설정
        msg.setText(arg_2)                          # 내용설정
        msg.exec_()       

if __name__ == "__main__" :
    app = QApplication(sys.argv)
    try:
        myWindow = MainWindow()
        myWindow.show()
        app.exec_()
    except Exception as e:
        msg = QMessageBox()
        msg.setWindowTitle("Error")               # 제목설정
        msg.setText(str(e))                          # 내용설정
        msg.exec_()  