import os
import sys
# import warnings
# import time

from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, QDate, QSize
from PyQt5 import uic
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

# 절대경로를 상대경로로 변경 하는 함수
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

#UI파일 연결
total_overtime= uic.loadUiType(resource_path("./ui/total_prod_order.ui"))[0] # Window 사용시 ui 주소
# main_window= uic.loadUiType(resource_path("/Users/black/projects/make_erp/main_window.ui"))[0] # Mac 사용시 ui 주소

#화면을 띄우는데 사용되는 Class 선언
class MainWindow(QWidget, total_overtime) :
    def __init__(self) :
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("생산오더 조회")

        self.layout_setting()
        self.slots()


    def layout_setting(self):
        # 버튼 레이아웃
        items_layout = QHBoxLayout()
        items_layout.addWidget(self.label)
        items_layout.addWidget(self.date_select_from)
        items_layout.addWidget(self.date_select_to)
        items_layout.addWidget(self.label_3)
        items_layout.addWidget(self.txt_prod_id)
        items_layout.addWidget(self.label_4)
        items_layout.addWidget(self.txt_item_id)
        items_layout.addWidget(self.label_5)
        items_layout.addWidget(self.txt_item_name)
        items_layout.addWidget(self.label_6)
        items_layout.addWidget(self.comb_prod_status)
        items_layout.addWidget(self.label_7)
        items_layout.addWidget(self.txt_sales_id)        
        items_layout.addWidget(self.btn_search)

        # 실행 버튼 레이아웃
        exec_layout = QHBoxLayout()
        exec_layout.addWidget(self.btn_download)
        exec_layout.setAlignment(Qt.AlignLeft)  # 왼쪽 정렬 추가

        # 전체 레이아웃
        main_layout = QVBoxLayout()
        main_layout.addLayout(items_layout)  # 버튼 추가
        main_layout.addWidget(self.tbl_info)  # 테이블 추가
        main_layout.addLayout(exec_layout)

        self.setLayout(main_layout)

        # 현재시간 설정
        self.set_date()

    def slots(self):
        self.btn_search.clicked.connect(self.get_args)
        self.btn_download.clicked.connect(self.make_file)
    #     self.btn_search_dept.clicked.connect(self.popup_dept_info)
    #     self.btn_clear.clicked.connect(self.clear)
    #     self.btn_close.clicked.connect(self.close)
    #     # self.btn_close.clicked.connect(self.window_close)
        # self.btn_select_emp.clicked.connect(self.popup_emp_info)

    def set_date(self):
        self.date_select_from.setDate(QDate.currentDate())
        self.date_select_to.setDate(QDate.currentDate())

    def clear(self):        
        self.tbl_info.setRowCount(0) # clear()는 행은 그대로 내용만 삭제, 행을 "0" 호출 한다.

        self.txt_dept_id.setText("")
        self.txt_dept_name.setText("")

    def get_args(self):
        #  sql문에 조건 검색을 위한 "%%" 처리
        item_id  = self.txt_item_id.text()
        if item_id == "":
            item_id = '%%'
        else:
            item_id = item_id

        item_name = self.txt_item_name.text()
        if item_name == "":
            item_name = '%%'
        else:
            item_name = item_name

        p_order_id = self.txt_prod_id.text()
        if p_order_id == "":
            p_order_id = '%%'
        else:
            p_order_id = p_order_id

        status = self.comb_prod_status.currentText()
        if status == "":
            status = '%%'
        else:
            status = status

        s_order_id = self.txt_sales_id.text()
        if s_order_id == "":
            s_order_id = '%%'
        else:
            s_order_id = s_order_id

        from_date = self.date_select_from.date().toString("yyyy-MM-dd")
        to_date = self.date_select_to.date().toString("yyyy-MM-dd")

        arr_1 = [from_date, to_date, p_order_id, item_id, item_name, status, s_order_id]
        self.make_data(arr_1)

    def make_data(self, arr_1):
           
        from db.db_select import Select
        select = Select()

        try:
            result, column_names = select.select_prod_info(arr_1)
            self.make_table(len(result), result, column_names)
        except Exception as e:
                print(e)
                self.msg_box("Error", str(e))

    def make_table(self, num, arr_1, column_names):   
        self.tbl_info.setRowCount(0) # clear()는 행은 그대로 내용만 삭제, 행을 "0" 호출 한다.

        col = len(column_names)

        self.tbl_info.setRowCount(num)
        self.tbl_info.setColumnCount(col)
        self.tbl_info.setHorizontalHeaderLabels(column_names)

        for i in range(num):
            for j in range(col):
                cell_value = arr_1[i][j]

                # NULL(None)을 공란으로 처리
                if cell_value is None:
                    cell_value = ""

                item = QTableWidgetItem(str(cell_value))
                self.tbl_info.setItem(i, j, item)

                # 7번째 컬럼(인덱스 6)의 숫자를 시간 형식으로 변환
                if j in [7,8] and cell_value != "":  # 8번째 컬럼
                    cell_value = self.format_seconds_to_time(cell_value)

                # 3번째 컬럼만 왼쪽 정렬
                if j == 2:  # 컬럼 인덱스 2 (3번째 컬럼)
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                else:  # 나머지 컬럼은 중앙 정렬
                    item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)                
                
                # self.tbl_info.setItem(i, j, QTableWidgetItem(str(arr_1[i][j])))
                
                # 전체 중앙 정렬
                #self.tbl_info.item(i, j).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)     

        # 테이블 서식 설정

        # 컬럼 데이터 정렬
        # 3번째 컬럼만 왼쪽 정렬
        # 컬럼 헤더를 인터랙티브 모드로 설정 + 데이터 길이에 맞추어 확장 가능하도록 설정
        table = self.tbl_info
        header = table.horizontalHeader()

        # 컬럼별 설정: 일부는 Interactive, 일부는 ResizeToContents
        for i in range(table.columnCount()):
            if i in [2, 5, 6]:  # 특정 컬럼은 길이에 맞추어 조정
                header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
                
            else:  # 나머지 컬럼은 Interactive
                header.setSectionResizeMode(i, QHeaderView.Interactive)

        # 정렬 기능 활성화
        self.tbl_info.setSortingEnabled(True)

        # 마지막 컬럼도 Stretch 비율로 포함
        header.setStretchLastSection(False)

        # 컬럼 헤더를 인터랙티브 모드로 설정하여 마우스로 조절 가능하게 함
        # header.setSectionResizeMode(QHeaderView.Interactive)

        # header.setStyleSheet("QHeaderView::section {"
        #              "background-color: #A9A9A9;"
        #              "color: black;"
        #             #  "font-weight: bold;"
        #              "font-size: 12px;"
        #              "padding: 5px;"
        #              "border: 1px solid #ccc; }")
        #             #  "border: 1px solid #ccc; }")
        
        # # 일정한 비율로 지정된 컬럼의 너비를 조정
        # ################################################################
        # table = self.tbl_info

        # # 테이블 위젯의 현재 너비 가져오기
        # total_width = table.viewport().width()

        # table.setColumnWidth(0, int(total_width * 1 / 10))
        # table.setColumnWidth(1, int(total_width * 1 / 10))
        # table.setColumnWidth(2, int(total_width * 2 / 10))
        # table.setColumnWidth(3, int(total_width * 1 / 10))
        # table.setColumnWidth(4, int(total_width * 1 / 10))
        # table.setColumnWidth(5, int(total_width * 1 / 10))
        # ################################################################

        #     # 컬럼 길이를 테이블 전체 너비의 비율로 설정
        # ################################################################
        # table = self.tbl_info
        # header = table.horizontalHeader()

        # # 각 컬럼의 비율 설정 (예: [2, 1, 1, 3, 3])
        # column_ratios = [1, 1, 3, 1, 1, 3]  # 비율은 원하는 대로 조정 가능
        # total_ratio = sum(column_ratios)

        # for i in range(col):
        #     stretch_factor = column_ratios[i] / total_ratio  # 각 컬럼 비율
        #     header.setSectionResizeMode(i, QHeaderView.Stretch)  # Stretch 모드 활성화
        #     table.setColumnWidth(i, int(table.width() * stretch_factor))  # 비율 기반 너비 설정
        # ################################################################

        # # 마지막 컬럼도 Stretch 비율로 포함
        # header.setStretchLastSection(True)

        # # 컨텐츠의 길이에 맞추어 컬럼의 길이를 자동으로 조절
        # # 이 방법은 컬럼의 길이를 마우스로 조절할 수 없게 함.
        # ################################################################
        # table = self.tbl_info
        # header = table.horizontalHeader()

        # for i in range(col):
        #     header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        # ################################################################
       


        # # 테이블의 길이에 맞추어 컬럼 길이를 균등하게 확장
        # # 마우스를 이용한 컬럼길이 조절 불가
        # self.tbl_info.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)




        
    # 테이블 선택범위 삭제
    def delete_rows(self):
        indexes = []
        rows = []

        for idx in self.tbl_info.selectedItems():
            indexes.append(idx.row())

        for value in indexes:
            if value not in rows:
                rows.append(value)

        # 삭제시 오류 방지를 위해 아래서 부터 삭제(리버스 소팅)
        rows = sorted(rows, reverse=True)

        # 선택행 삭제
        for rowid in rows:
            self.tbl_info.removeRow(rowid)

    # 부서명 가져오기 팝업
    def popup_dept_info(self):
        from popup.dept_popup import DeptWindow
        input_dialog = DeptWindow()

        if input_dialog.exec_():
            value = input_dialog.get_input_value()

            try:
                self.txt_dept_id.setText(value[0].text())
                self.txt_dept_name.setText(value[1].text())
            except:
                return
        
    ### 다이알로그 창으로 값을 전달 할 때는 아규먼트를 보내 주는 방식으로 !!!!
    # def popup_emp_info(self):
    #     arg_1 = self.txt_dept_id.toPlainText()
    #     input_dialog = EmpWindow(arg_1) ##   <-----중요 포인트
    #     if input_dialog.exec_():
    #         value = input_dialog.get_input_value()

    #     try:
    #         self.txt_emp_id.setText(value[2].text())
    #         self.txt_emp_name.setText(value[3].text())
    #     except:
    #         return
        
    def get_dept_id(self):
        print(self.dept_id)
        return self.dept_id
    
      # 테이블에 남겨진 정보를 엑셀로 변환
    def make_file(self):
        rows = self.tbl_info.rowCount()
        cols = self.tbl_info.columnCount()

        headers = []
        for i in range(cols):
            headers.append(self.tbl_info.horizontalHeaderItem(i).text())

        list_2 = [] # 최종적으로 사용할 리스트는 for문 밖에 선언
        for i in range(rows):
            list_1 = [] # 2번째 for문 안쪽에서 사용할 리스트 선언
            for j in range(cols): 
                data = self.tbl_info.item(i,j)
                list_1.append(data.text())
            list_2.append(list_1)

        num = len(list_2)
        self.make_excel(list_2, num, headers)
        

    # 엑셀 파일을 만들고 넘겨진 배열 정보를 이용하여 sheet에 정보를 기입/저장 함.
    def make_excel(self, arr, num, headers):
        wb = openpyxl.Workbook()
        wb.create_sheet(index=0, title='작업진행현황')

        sheet = wb.active
        sheet.append(headers)

        for i in range(num):
            for j in range(len(headers)):
                sheet.cell(row=i+2, column=j+1, value=arr[i][j])

        ## 각 칼럼에 대해서 모든 셀값의 문자열 개수에서 1.1만큼 곱한 것들 중 최대값을 계산한다.
        for column_cells in sheet.columns:
            # length = max(len(str(cell.value))*1.1 for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = 20
            ## 셀 가운데 정렬
            for cell in sheet[column_cells[0].column_letter]:
                cell.alignment = Alignment(horizontal='center')
        
        fname = self.file_save()

        try:
            if fname:
                self.save_excel(wb, fname)
        except Exception as e:
            self.msg_box("Error", str(e))


    # 파일 저장 대화상자(파일명 만들기)
    def file_save(self):
        now = datetime.now()
        arg_1 = now.strftime('%Y-%m-%d %H-%M-%S')
        adress = "./excel/download_" + arg_1 + ".xlsx"

        dialog = QFileDialog(self)
        qurl  = dialog.getSaveFileName(parent=self, caption='Save file', directory=adress)
        
        url = qurl[0]
        try:
            return url
        except Exception as e:
            QMessageBox.about(self, 'Warning', e)


    def save_excel(self, workbook, file_name):
        workbook.save(file_name)


    def format_seconds_to_time(self, seconds):
        """초(sec)를 '00시 00분 00초' 형식으로 변환"""
        try:
            seconds = int(seconds)  # 초를 정수로 변환
            hours = seconds // 3600
            minutes = (seconds % 3600) // 60
            secs = seconds % 60
            return f"{hours:02d}시 {minutes:02d}분 {secs:02d}초"
        except ValueError:
            # 변환할 수 없는 값은 그대로 반환
            return seconds

    def msg_box(self, arg_1, arg_2):
        msg = QMessageBox()
        msg.setWindowTitle(arg_1)               # 제목설정
        msg.setText(arg_2)                          # 내용설정
        msg.exec_()                                 # 메세지박스 실행

    def window_close(self):
        self.close()

if __name__ == "__main__" :
    app = QApplication(sys.argv)
    myWindow = MainWindow()
    myWindow.show()
    app.exec_()